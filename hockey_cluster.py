#SENG474 Hockey Prediction
from collections import Counter 
import numpy as np
import openpyxl as pyxl
import math

#Code for getting the csv spreadsheet. Adjust the filepath according to your needs.
filepath = "C:\Users\Keith\Desktop\School\DataMining\HockeyReference2009-2013.xlsx"
wb = pyxl.load_workbook(filepath)
ws = wb.get_sheet_by_name('HockeyReference2009-2013')

#This will be a list of candidate classes. We can refer to this later when
#for training and testing purposes.
candidate_list = []
stats_matrix = []
"""
gp = index 0
g = index 1
a = index 2
pts = index 3
plus/minus = index 4
pim_list = index 5
ev_list = index 6
pp_list = index 7
sh_list = index 8
gw = index 9
ev_list 10
pp_list 11
sh_list 12
s_list = index 13
spercentage_list = index 4
toi_list = index 15
atoi_list = index 16
blk_list = index 17
hit_list = index 18
fow_list = index 19
fol_list = index 20
fopercentage_list = index 21
"""
#player class. These are the players we will be comparing to player 1
class candidates():
	def __init__(self, name, height, pointspergame, age):
		self._name = name
		self._height = height
		self._pointspergame = pointspergame
		self._age = age

#This function looks for the specified player in the csv spreadsheet.
def findplayer(player, start_row, end_row):
	global p1_name
	global p1_age
	global p1_height
	global p1_pointspergame
	global p1_row
	#Iterate through each row in column B to find matching player name.
	#max row can be adjusted according to the number of players.
	for row in ws.iter_rows(min_row = start_row, max_row = end_row, min_col = 2, max_col = 2):
		for cell in row: #for each cell in the current row
			s = str(cell.value) #need to cast to string because default value of cell.value is unicode.
			
			#String splitting to filter out \ and *.
			s = string_split(s)
			
			#Assign values to player 1's attributes if we can find them.
			if s == player:
				p1_name = s
				print "Player: %s" % p1_name
				p1_age = ws.cell(row = cell.row, column = 3).value
				print "Age: %d" % p1_age
				p1_height = ws.cell(row = cell.row, column = 28).value
				print "Height: %dcm" % p1_height
				p1_pointspergame = float(ws.cell(row = cell.row, column = 9).value) / float(ws.cell(row= cell.row, column = 6).value)
				print "Points per game: %f" % p1_pointspergame
				p1_row = cell.row
				print "row: %d" % p1_row
				return True
	return False

#This function finds potential candidates similar to player 1.
def find_matching_players(player, start_row, end_row):
	p2_name = ""
	p2_height = 0
	p2_age = 0
	p2_pointspergame = 0
	max_thresh = [0, 0, 0]
	min_thresh = [200000, 200000, 200000]
	
	for x in range(start_row, end_row+1):
		if x != p1_row: #We don't want to match player 1 with themself.
			s = str(ws.cell(row = x, column = 2).value)
			p2_name = string_split(s) #More String splitting to filter out * and \
			p2_age = ws.cell(row = x, column = 3).value
			p2_height = ws.cell(row = x, column = 28).value
			p2_pointspergame = float(ws.cell(row = x, column = 9).value) / float(ws.cell(row = x, column = 6).value)
			threshold = [0, 0, 0]
			"""
			formula
			threshold 0 = sqrt((points per game P1-points per game P2)^2)/points per game P1
			threshold 1 = sqrt((height p1 - height p2)^2)/ height p1
			threshold 2 = sqrt((age p1 - age p2)^2)/age p1
			requirements for candidate consideration:
			1 - threshold
			minimum games played = 10
			
			"""
			threshold[0] = (math.sqrt((p1_pointspergame-p2_pointspergame)**2)/p1_pointspergame)
			threshold[1] = (math.sqrt((p1_height-p2_height)**2)/p1_height)
			threshold[2] = (math.sqrt((p1_age-p2_age)**2)/p1_age)
			#threshold = math.sqrt(((p1_pointspergame-p2_pointspergame)**2) + ((p1_height-p2_height)**2) + ((p1_age-p2_age)**2)) / math.sqrt(0.95*(p1_pointspergame**2) + 0.025*(p1_height**2) + 0.025*(p1_age**2))
			#add them to candidate_list
			for i in range(0, len(threshold)):
				if max_thresh[i] < threshold[i]:
					max_thresh[i] = threshold[i]
				if min_thresh[i] > threshold[i]:
					min_thresh[i] = threshold[i]
			if threshold[0] <= 0.1 and threshold[1] <= 0.1 and threshold[2] <= 0.1 and ws.cell(row = x, column = 6).value >= 10:
				p2 = candidates(p2_name, p2_height, p2_pointspergame, p2_age)
				candidate_list.append(p2)
	for t in range(0, 3):	
		if t == 0:
			print "Max threshold of points per game: "
		if t == 1:
			print "Max threshold of height: "
		if t == 2:
			print "Max threshold of age: "
		print "%f" % max_thresh[t]
		if t == 0:
			print "Min threshold of points per game: "
		if t == 1:
			print "Min threshold of height: "
		if t == 2:
			print "Min threshold of age: "
		print "%f" % min_thresh[t]
	print "Candidate list size: %d" % len(candidate_list)
	"""
	for c in candidate_list:
		print c._name
		print "age: %d" % c._age
		print "height: %d" % c._height
		print "points per game: %f" % c._pointspergame
	"""
	
	return
	
def find_next_year_stats(start_row, end_row):
	for i in range(0, 22):
		stats_matrix.append([])
		
	for c in candidate_list:
		for r in range(start_row, end_row+1):
			s = str(ws.cell(row = r, column = 2).value)
			#String splitting to filter out \ and *.
			s = string_split(s)
			if s == c._name:
				for col in range(6, 28):
					stat = ws.cell(row = r, column = col).value
					stats_matrix[col - 6].append(stat)
				break
	#print "Number of players still playing in next year: %d" % len(stats_matrix[0])
	return

#predict the player's points and points per game for next year
def predict_results(name):
	print "Prediction for %s:" % name
	"""
	freq_points = Counter(stats_matrix[3]).most_common(1)
	freq_gamesplayed = Counter(stats_matrix[0]).most_common(1)
	freq_pointspergame = float(freq_points[0][0])/float(freq_gamesplayed[0][0])
	"""
	freq_points = np.mean(stats_matrix[3])
	freq_gamesplayed = np.mean(stats_matrix[0])
	freq_pointspergame = float(freq_points)/float(freq_gamesplayed)
	print "Points: %d " % freq_points
	print "Points per game: %f" % freq_pointspergame
	"""
	for x in range(0, 22):
		stat = Counter(stats_matrix[x])
		frequent_stat = stat.most_common(1)
		current_stat = str(ws.cell(row = 2, column = x+6).value)
		print "%s: " % current_stat
		print frequent_stat
	"""
	return

def actual_results(name, start_row, end_row):
	print "Actual results for %s: " % name
	for r in range(start_row, end_row+1):
		s = str(ws.cell(row = r, column = 2).value)
		#String splitting to filter out \ and *.
		s = string_split(s)
		if s == name:
			print "Points: %d" % ws.cell(row = r, column = 9).value
			ppg = float(ws.cell(row = r, column = 9).value) / float(ws.cell(row = r, column = 6).value)
			print "Points per game: %f" % ppg
			break
	return

#string split to filter out * and \
def string_split(s):
	if "*" in s:
		return s.split('*')[0]
	return s.split('\\')[0]
"""
formula
threshold 1 = sqrt((points per game P1-points per game P2)^2)/points per game P1
threshold 2 = sqrt((height p1 - height p2)^2)/ height p1
threshold 3 = sqrt((age p1 - age p2)^2)/age p1
requirements for candidate consideration:
minimum threshold = 0.9	for each threshold	
minimum games played = 10
"""		

"""
Make numpy arrays from spreadsheet
print(wb.get_sheet_names())
ws = wb['HockeyReference2009.csv']
print(ws['E1'].value)
table = np.array([[cell.value for cell in col] for col in ws['B1':'B881']])
print(table)
"""
"""
Rows 3 to 881 = 2009
Rows 882 to 1772 = 2010
Rows 1773 to 2663 = 2011
Rows 2664 to 3502 = 2012
Rows 3503 to 4388 = 2013
"""
#Calling code
name = input('Enter player name as: "player name" ')
if findplayer(name, 3, 881):
	find_matching_players(name, 3, 881)
	"""
	for c in candidate_list:
		print c._name
		print c._age
		print c._height
		print c._pointspergame
	"""
	find_next_year_stats(882, 1772)
	predict_results(name)
	actual_results(name, 882, 1772)
else:
	print "Player not found"