from collections import Counter 
import numpy as np
import openpyxl as pyxl
import math
from sklearn.neural_network import MLPClassifier

#Code for getting the csv spreadsheet. Adjust the filepath according to your needs.
filepath = "C:\Users\Keith\Desktop\School\DataMining\HockeyReference2009-2013.xlsx"
wb = pyxl.load_workbook(filepath)
ws = wb.get_sheet_by_name('HockeyReference2009-2013')

data = []
lookup = []
classes = []
predict = []
average = []

"""
1 = Player
2 = Age
3 = Games Played
4 = Points
5 = Height
"""

#player class. These are the players we will be comparing to player 1
class player():
	def __init__(self, name, height, pointspergame, age):
		self._name = name
		self._height = height
		self._pointspergame = pointspergame
		self._age = age

#This function looks for the specified player in the csv spreadsheet.
def populateData(start_row, end_row):
	#Iterate through each row in column A to find matching player name.
	#max row can be adjusted according to the number of players.
	for row in ws.iter_rows(min_row = start_row, max_row = end_row):
		for cell in row: #for each cell in the current row
			p = []
			look = []
			name = str(cell.value.encode('utf-8')) #need to cast to string because default value of cell.value is unicode.
			
			#String splitting to filter out \ and *.
			name = string_split(name)
			name2 = str(ws.cell(row = cell.row+1, column = 1).value.encode('utf-8'))
			name2 = string_split(name2)
			
			age = 0.037*(float(ws.cell(row = cell.row, column = 2).value) - 17)
			ppg = float(ws.cell(row = cell.row, column = 4).value) / float(ws.cell(row = cell.row, column = 3).value)
			height = 0.0238*(float(ws.cell(row = cell.row, column = 5).value) - 164)
			#toi = 
			if (name == name2):
				lookForward = float(ws.cell(row = cell.row+1, column = 4).value) / float(ws.cell(row = cell.row+1, column = 3).value)
			else:
				break

			points = ppg * 82
			points = int(round(points))
			classes.append(points)
			p.append(ppg)
			p.append(age)
			p.append(height)

			data.append(p)

			look.append(points)
			look.append(ppg)
			look.append(age)
			look.append(height)
			look.append(name)
			look.append(lookForward)

			lookup.append(look)
			break
	return True

#string split to filter out * and \
def string_split(s):
	if "*" in s:
		return s.split('*')[0]
	return s.split('\\')[0]

ppgIN = input('Enter player points per game:')
ageIN = input('Enter player age in years (18-44):')
heightIN = input('Enter player height in cm (165-205):')
"""
ppgIN = 0.791667
ageIN = 29
heightIN = 180
"""
ageIN = 0.037*(ageIN - 17)
heightIN = 0.0238*(heightIN - 164)

populateData(2, 4900)
i = 0
"""
while i < len(data):
	classes.append(i)
	i = i+1
"""

clf = MLPClassifier(solver='sgd', hidden_layer_sizes=(200, 200, 200), max_iter=3000, tol=0.0001)

clf.fit(data, classes)

#data(ppg, age, height)
#predict = clf.predict([[0.48148, 0.037, 0.2618]])
#predict = clf.predict([[0,0,0],[1,1,1]])
predict = clf.predict([[ppgIN, ageIN, heightIN]])

#lookup(points, ppg, age, height, name, lookForward)
j = 0
k = 0
while j < len(predict):
	average = []
	while k < len(lookup):
		if lookup[k][0] == predict[j]:
			average.append(lookup[k][5])
		k = k+1
	print sum(average)/len(average)
	j = j+1


#lookup(ppg, age, height, name, lookForward)
"""
j = 0
while j < len(predict):
	print "Name: ", lookup[predict[j]][3]
	print "ppg: ", lookup[predict[j]][0]
	print "age: ", ((lookup[predict[j]][1])/0.037)+17
	print "height: ", ((lookup[predict[j]][1])/0.0238)+164
	print "Next season's predicted points: ", lookup[predict[j]][4]
	j = j+1
"""